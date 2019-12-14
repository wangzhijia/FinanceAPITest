from collections import namedtuple

from openpyxl import load_workbook
from scripts.handle_config import do_config
from scripts.constants import TEST_DATAS_FILE_PATH


class HandleExcel(object):
    """
    定义处理Excel的类
    """

    def __init__(self, filename, sheetname=None):
        self.filename = filename
        self.sheetname = sheetname
        # 打开Excel文件
        self.wb = load_workbook(self.filename)
        # 定位表单
        self.ws = self.wb[self.sheetname] if self.sheetname is not None else self.wb.active

        self.sheet_head_tuple = tuple(self.ws.iter_rows(max_row=1, values_only=True))[0]
        # 创建一个元组类
        self.Cases = namedtuple("Cases",self.sheet_head_tuple)
        # 定义一个存放所有Cases命名元组对象
        self.cases_list = []

    def get_cases(self):
        """
        获取所有测试用例
        :return:
        """
        for data in self.ws.iter_rows(min_row=2, values_only=True):
            self.cases_list.append(self.Cases(*data))

        return self.cases_list

    def get_case(self, row):
        """
        获取某一条测试数据
        :param row:
        :return:
        """
        if isinstance(row, int) and ( 1 <= row <= self.ws.max_row):
            return tuple(self.ws.iter_rows(min_row=row, max_row=row, values_only=True))[0]
        else:
            print("传入行号有误,行号应是大于等于1的整数")

    def write_result(self, row, actual, result):
        """
        将实际值与测试用例执行的结果保存到Excel中
        :param row:
        :param actual:
        :param ruselt:
        :return:
        """
        other_wb = load_workbook(self.filename)
        other_ws = other_wb[self.sheetname]

        if isinstance(row, int) and (2 <= row <= other_ws.max_row):
            other_ws.cell(row=row, column=do_config("excel","actual_col"), value=actual)
            other_ws.cell(row=row, column=do_config("excel","result_col"), value=result)
            other_wb.save(self.filename)
        else:
            print("传入行号有误,行号应是大于等于2的整数")


if __name__ == '__main__':
    one_excel = HandleExcel(filename=TEST_DATAS_FILE_PATH)
    cases = one_excel.get_cases()
    print(cases)